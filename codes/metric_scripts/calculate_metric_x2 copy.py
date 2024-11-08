import os
import math
import numpy as np
import cv2
import glob
from sewar.full_ref import sam, uqi, scc
from openpyxl import Workbook, load_workbook
import torch
import lpips  # Import the LPIPS library


def main():
    # Configurations
    folder_GT = r'D:\Wz_Project\KanFVRSR-S2\dataset\Sentinel2\Sentinel2_b2\val\HR'
    folder_Gen = r'D:\Wz_Project\KanFVRSR-S2\experiment\FVRSRx2_Sentinel2_b2\results'

    crop_border = 4
    suffix = ''
    test_Y = False

    PSNR_all = []
    SSIM_all = []
    SAM_all = []
    QI_all = []
    SCC_all = []
    LPIPS_all = []  # New list for storing LPIPS values

    # Initialize LPIPS model (use 'alex', 'vgg', or 'squeeze')
    lpips_model = lpips.LPIPS(net='alex').to('cuda')  # Move model to GPU if available

    img_list = sorted(glob.glob(os.path.join(folder_GT, '*')))

    if test_Y:
        print('Testing Y channel.')
    else:
        print('Testing RGB channels.')

    results = []

    for i, img_path in enumerate(img_list):
        base_name = os.path.splitext(os.path.basename(img_path))[0]
        im_GT = cv2.imread(img_path)
        
        # Check if im_GT was loaded correctly
        if im_GT is None:
            print(f"Warning: Failed to load image {img_path}. Skipping.")
            continue
        
        im_GT = im_GT / 255.
        im_Gen = cv2.imread(os.path.join(folder_Gen, base_name + suffix + '.tif'))
        
        # Check if im_Gen was loaded correctly
        if im_Gen is None:
            print(f"Warning: Failed to load generated image for {base_name}. Skipping.")
            continue
        
        im_Gen = im_Gen / 255.

        if test_Y and im_GT.shape[2] == 3:
            im_GT_in = bgr2ycbcr(im_GT)
            im_Gen_in = bgr2ycbcr(im_Gen)
        else:
            im_GT_in = im_GT
            im_Gen_in = im_Gen

        if crop_border > 0:
            if im_GT_in.ndim == 3:
                cropped_GT = im_GT_in[crop_border:-crop_border, crop_border:-crop_border, :]
                cropped_Gen = im_Gen_in[crop_border:-crop_border, crop_border:-crop_border, :]
            elif im_GT_in.ndim == 2:
                cropped_GT = im_GT_in[crop_border:-crop_border, crop_border:-crop_border]
                cropped_Gen = im_Gen_in[crop_border:-crop_border, crop_border:-crop_border]
            else:
                raise ValueError(f'Wrong image dimension: {im_GT_in.ndim}. Should be 2 or 3.')
        else:
            cropped_GT = im_GT_in
            cropped_Gen = im_Gen_in

        # Convert images to torch tensors for LPIPS
        im_GT_tensor = torch.tensor(cropped_GT).permute(2, 0, 1).unsqueeze(0).float().to('cuda') * 2 - 1
        im_Gen_tensor = torch.tensor(cropped_Gen).permute(2, 0, 1).unsqueeze(0).float().to('cuda') * 2 - 1

        # LPIPS score calculation
        LPIPS_value = lpips_model(im_GT_tensor, im_Gen_tensor).item()

        PSNR = calculate_rgb_psnr(cropped_GT * 255, cropped_Gen * 255)
        SSIM = calculate_ssim(cropped_GT * 255, cropped_Gen * 255)
        SAM = sam(cropped_GT * 255, cropped_Gen * 255)
        QI = uqi(cropped_GT * 255, cropped_Gen * 255)
        SCC = scc(cropped_GT * 255, cropped_Gen * 255)

        print(f'{i + 1:3d} - {base_name:25}. \tPSNR: {PSNR:.6f} dB, \tSSIM: {SSIM:.6f}, \tSAM: {SAM:.6f}, \tQI: {QI:.6f}, \tSCC: {SCC:.6f}, \tLPIPS: {LPIPS_value:.6f}')
        
        PSNR_all.append(PSNR)
        SSIM_all.append(SSIM)
        SAM_all.append(SAM)
        QI_all.append(QI)
        SCC_all.append(SCC)
        LPIPS_all.append(LPIPS_value)  # Store LPIPS values
        
        results.append([base_name, PSNR, SSIM, SAM, QI, SCC, LPIPS_value])

    print('Average: PSNR: {:.6f} dB, SSIM: {:.6f}, SCC: {:.6f}, SAM: {:.6f}, QI: {:.6f}, LPIPS: {:.6f}'.format(
        sum(PSNR_all) / len(PSNR_all),
        sum(SSIM_all) / len(SSIM_all),
        sum(SCC_all) / len(SCC_all),
        sum(SAM_all) / len(SAM_all),
        sum(QI_all) / len(QI_all),
        sum(LPIPS_all) / len(LPIPS_all),  
    ))

    file_path = r'D:\Wz_Project\KanFVRSR-S2\experiment\FVRSRx2_Sentinel2_b2\results\x2_Sentinel2.xlsx'
    save_data_to_excel(results, file_path)

# Rest of the code...


def bgr2ycbcr(img, only_y=True):
    in_img_type = img.dtype
    img.astype(np.float32)
    if in_img_type != np.uint8:
        img *= 255.
    if only_y:
        rlt = np.dot(img, [24.966, 128.553, 65.481]) / 255.0 + 16.0
    else:
        rlt = np.matmul(img, [[24.966, 112.0, -18.214], [128.553, -74.203, -93.786],
                              [65.481, -37.797, 112.0]]) / 255.0 + [16, 128, 128]
    if in_img_type == np.uint8:
        rlt = rlt.round()
    else:
        rlt /= 255.
    return rlt.astype(in_img_type)

def save_data_to_excel(results, file_name):
    try:
        try:
            wb = load_workbook(filename=file_name)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["File Name", "PSNR", "SSIM", "SAM", "QI", "SCC", "LPIPS"])  # Add LPIPS header

        for result in results:
            ws.append(result)

        wb.save(file_name)
        print("Data saved to Excel file:", file_name)
    except Exception as e:
        print("Error saving data to Excel:", str(e))

def calculate_psnr(img1, img2):
    img1 = img1.astype(np.float64)
    img2 = img2.astype(np.float64)
    mse = np.mean((img1 - img2) ** 2)
    if mse == 0:
        return float('inf')
    return 20 * math.log10(255.0 / math.sqrt(mse))

def calculate_rgb_psnr(img1, img2):
    n_channels = img1.shape[2]
    sum_psnr = 0
    for i in range(n_channels):
        this_psnr = calculate_psnr(img1[:, :, i], img2[:, :, i])
        sum_psnr += this_psnr
    return sum_psnr / n_channels

def ssim(img1, img2):
    C1 = (0.01 * 255) ** 2
    C2 = (0.03 * 255) ** 2

    img1 = img1.astype(np.float64)
    img2 = img2.astype(np.float64)
    kernel = cv2.getGaussianKernel(11, 1.5)
    window = np.outer(kernel, kernel.transpose())

    mu1 = cv2.filter2D(img1, -1, window)[5:-5, 5:-5]
    mu2 = cv2.filter2D(img2, -1, window)[5:-5, 5:-5]
    mu1_sq = mu1 ** 2
    mu2_sq = mu2 ** 2
    mu1_mu2 = mu1 * mu2
    sigma1_sq = cv2.filter2D(img1 ** 2, -1, window)[5:-5, 5:-5] - mu1_sq
    sigma2_sq = cv2.filter2D(img2 ** 2, -1, window)[5:-5, 5:-5] - mu2_sq
    sigma12 = cv2.filter2D(img1 * img2, -1, window)[5:-5, 5:-5] - mu1_mu2

    ssim_map = ((2 * mu1_mu2 + C1) * (2 * sigma12 + C2)) / ((mu1_sq + mu2_sq + C1) *
                                                            (sigma1_sq + sigma2_sq + C2))
    return ssim_map.mean()

def calculate_ssim(img1, img2):
    if not img1.shape == img2.shape:
        raise ValueError('Input images must have the same dimensions.')
    if img1.ndim == 2:
        return ssim(img1, img2)
    elif img1.ndim == 3:
        if img1.shape[2] == 3:
            ssims = []
            for i in range(img1.shape[2]):
                ssims.append(ssim(img1[..., i], img2[..., i]))
            return np.array(ssims).mean()
        elif img1.shape[2] == 1:
            return ssim(np.squeeze(img1), np.squeeze(img2))
    else:
        raise ValueError('Wrong input image dimensions.')

if __name__ == '__main__':
    main()
